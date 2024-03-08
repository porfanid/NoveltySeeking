from get_user_answers_from_database import get_answers
from get_questions import get_questions
import openpyxl

file_path = "./excel.xlsx"
sheet_name = "NoveltySeeking"

workbook = openpyxl.Workbook()
# Select the sheet by name
sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
sheet.title = sheet_name

def write_to_excel(data, first_line=False):
    
    # Find the next available row
    if first_line:
        row_number = sheet.max_row
    else:
        row_number = sheet.max_row + 1
    # Write data to the row
    for col_num, value in enumerate(data, start=1):
        sheet.cell(row=row_number, column=col_num, value=value)

column_titles = ["Κωδικός", "Γενικά αποτελέσματα", "Πόσες ερωτήσεις απάντησε σωστά", "Επιλογή κατηγορίας"]
for i in range(1,8):
    column_titles.append("Ερώτηση"+str(i))
    column_titles.append("Επιλογή για επόμενο βίντεο"+str(i))


write_to_excel(column_titles, True)

users = get_answers()
questions = get_questions()

number_for_correct_answer = 1
number_for_wrong_answer = 2

section_index = {
    "Animals": 1,
    "City": 2,
    "Ocean": 3,
    "Space": 4
}

for user, answers in users.items():
    previous_category = None
    previous_subcategory = None
    correct_answers = 0
    row = []
    count_final_choice = {
        1: 0, 2:0, 3:0
    }
    for index, answer in answers.items():
        if "choice" in answer:
            
            if index==1:
                category = section_index[answer["choice"]]
                row.append(category)
            else:
                final_choice = 1 # video of the same animal
                if answer["choice"]!=previous_category:
                    final_choice = 3 # video of another category
                elif answer["category"]!=previous_subcategory:
                    final_choice = 2 # video of another animal in the same category
                
                count_final_choice[final_choice]+=1
                row.append(final_choice)
            print(answer)
            has_answered_correctly=answer["quiz"]==questions[answer["choice"]][answer["category"]][answer["counter"]]["correct answer"]
            if has_answered_correctly:
                correct_answers+=1
            quiz_answer = number_for_correct_answer if has_answered_correctly else number_for_wrong_answer
            row.append(quiz_answer)
            previous_category =answer["choice"]
            previous_subcategory = answer["category"]
            
        else:
            pass

    user_data = [user, str(count_final_choice[1])+str(count_final_choice[2])+str(count_final_choice[3]), correct_answers]+row
    print(user_data)
    write_to_excel(user_data)
    
# Save the workbook
workbook.save(file_path)
print(f"Data written to {sheet_name} in {file_path}")